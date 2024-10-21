import csv
import os
import zipfile
from io import TextIOWrapper
from zipfile import ZipFile
from openpyxl import load_workbook
from pypdf import PdfReader

def test_create_archive():
    if not os.path.exists('../Resourses'): # проверяем существует ли папка
        os.mkdir('../Resourses') # создаем папку если её нет
    with zipfile.ZipFile('../Resourses/test_archive.zip', 'w') as zf: # создаем архив
        for file in ['Artem.xlsx', 'behaviors.csv', 'Safe.pdf']: # добавляем файлы в архив
            add_file = os.path.join('../tmp', file) # склеиваем путь к файлам которые добавляют в архив
            zf.write(add_file, os.path.basename(add_file)) # добавляем файл в архив


def test_csv():
    with zipfile.ZipFile('../Resourses/test_archive.zip') as zip_file: # открываем архив
        with zip_file.open('behaviors.csv') as csv_file: # открываем файл в архиве
            csvreader = list(csv.reader(TextIOWrapper(csv_file, 'utf-8-sig'))) # читаем содержимое файла и преобразуем его в список и декодируем его если в файле есть символы не из английского алфавита
            second_cell = csvreader[0] # получаем вторую строку - лист из одного элемента
            str = ';'.join(second_cell)  #разделяем строку на элементы
            second_row = str.split(';')
            assert second_row[0] == 'Feature' # проверка значения элемента в первом столбце второй строки
            assert second_row[1] == 'Story' # проверка значения элемента во втором столбце второй строки

def test_pdf():
    with zipfile.ZipFile('../Resourses/test_archive.zip') as zip_file:  # открываем архив
        with zip_file.open('Safe.pdf') as pdf_file:  # открываем файл в архиве
            reader = PdfReader(pdf_file)
            page = reader.pages[0].extract_text()
            assert 'Ижевск, 2024' in page

def test_xlsx():
    with zipfile.ZipFile('../Resourses/test_archive.zip') as zip_file:  # открываем архив
        with zip_file.open('Artem.xlsx') as xlsx_file:  # открываем файл в архиве
            workbook = load_workbook(xlsx_file)  # открываем файл
            sheet = workbook.active  # получаем активный лист
            assert 'Создание ПП' in sheet.cell(row=5, column=1).value
            assert 'Создание ОП, активация' in sheet.cell(row=6, column=1).value